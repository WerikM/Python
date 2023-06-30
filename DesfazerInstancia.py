import openpyxl
import time
import easygui
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from lxml import etree
import lxml.html

# Configura a instância do navegador
driver = webdriver.Edge()

# seleciona o arquivo da planilha
root = Tk()
root.withdraw()
file_path = askopenfilename()

# abrir a planilha
workbook = openpyxl.load_workbook(file_path)

# selecionar a planilha desejada
try:
    sheet = workbook['Plan1']
except KeyError:
    # caso não encontre, tentar selecionar a planilha com nome "Planilha1"
    try:
        sheet = workbook['Planilha1']
    except KeyError:
        # se não encontrar nenhuma das duas, mostra uma mensagem de erro e sai do programa
        print("Nenhuma planilha encontrada com os nomes 'Plan1' ou 'Planilha1'.")
        exit()

# Usuario colocara Login e Senha para o Safra
login = easygui.enterbox("Entre com o login:")
senha = easygui.passwordbox("Entre com a senha:")

# URL do Safra
url_safra = "SITE"

# Percorre as células da coluna A e realiza a navegação para cada valor
for a in range(2, sheet.max_row+1):
    # Obtém a pesquisa a partir da coluna A da linha atual
    termo_pesquisa = str(sheet.cell(row=a, column=1).value)
    print(f"Buscando por: {termo_pesquisa}")
    print(f"pesquisados: {a}")
    if sheet.cell(row=a, column=2).value is None:
        
      # Concatena com a URL do Safra
      url_pesquisa = url_safra + termo_pesquisa
      
      # Navega até o site do Safra
      driver.get(url_pesquisa)

      
      # Espera até que o campo de usuário esteja visível
      try:
          user_input = WebDriverWait(driver, 2).until(
              EC.visibility_of_element_located((By.CSS_SELECTOR, "input.b2w-user"))
          )
      except:
          # se o campo de usuário não estiver visível, significa que já foi feito o login,
          # então podemos pular para a próxima etapa
          pass
      else:
          # Escreve no campo de usuário    bruna.gaspere
          user_input.send_keys(login)
          
          # Encontra o campo de senha e escreve nele    Miguel*12
          password_input = driver.find_element(By.XPATH, "//input[@type='password']")
          password_input.send_keys(senha)
          
          # localiza o elemento select
          elemento_select = driver.find_element(By.NAME, "ad")
          
          # cria um objeto da classe Select a partir do elemento select
          select = Select(elemento_select)
          
          # seleciona a opção pelo valor
          select.select_by_value('AD')
          
          # Localiza o botão "Entrar" e clica nele
          botao_entrar = driver.find_element(By.XPATH, "//button[contains(text(), 'Entrar')]")
          botao_entrar.click()
  
      try:
          # Espera carregar a pagina por 1 segs
          time.sleep(2)
          # Espera até que o botão seja visível na página
          botao_desfazer_instancia = WebDriverWait(driver, 2).until(
              EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[2]/form/div/div[3]/div[2]/div[4]/button[9]")))
          botao_desfazer_instancia.click()
          elementos_supremo = driver.find_element(By.XPATH, "/html/body/div[1]")
                
          elementos_pai = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/form/div/div[3]/div[1]/div[1]")
  
      except:
          # Escreve "Falha" na coluna B se a verificação foi mal sucedida
          sheet.cell(row=a, column=2).value = "Sem desfazer"
          pass
          
      else:
          try:                       
              #antiga verificação do numero de entrega                                                                 
              # obter o HTML da página
              #html = driver.page_source
              #tree = lxml.html.fromstring(html)

              print(f"Entrega do site: {termo_pesquisa}")
  
  
          except:
              # Escreve "Falha 01" na coluna B se a verificação foi mal sucedida
              sheet.cell(row=a, column=2).value = "Falha 01"
              pass
              
          else: 
              
            # Localize todos os descendentes que contêm a palavra "Selecionar"
            i = 0
            Selecionar = elementos_pai.find_elements(By.XPATH, ".//*[contains(text(), 'Selecionar')]")
            print("Quantos 'Selecionar' tem? " + str(len(Selecionar)))
            
            for i, Selecionar in enumerate(Selecionar):
              Selecionar.click()
              
            motivo1 = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, ".//*[contains(text(), 'Informe aqui')]")))
            time.sleep(2)
            motivo1.click()
             # obter o HTML da página
            #html = driver.page_source
            #tree = lxml.html.fromstring(html)
             
             # encontrar o elemento pai do motivo1
            #motivo1_pai = driver.find_element(By.XPATH, "..")
        
            #motivo1_pai.click()
            motivo2 = elementos_pai.find_element(By.XPATH, ".//th")
            motivo2.click()
            try:
              motivo3 = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, ".//*[contains(text(), 'Cliente nao concordou com o prazo da devolucao')]")))
              motivo3.click()
            except:
              motivo3extra = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, ".//*[contains(text(), 'Arrependimento - DESISTÊNCIA DE COMPRA - Antes do faturamento - Engano na compra')]")))
              motivo3extra.click()
            time.sleep(1)
            motivo4 = elementos_pai.find_element(By.XPATH, ".//button")
            motivo4.click()
            texto_motivo = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, ".//textarea")))
            #texto_motivo = elementos_supremo.find_element(By.XPATH, ".//textarea")
            texto_motivo.send_keys("Cliente não concordou com o prazo da devolução")
            time.sleep(1)
            Confirmar_motivo = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, ".//*[contains(text(), ' Sim')]")))
            Confirmar_motivo.click()
            fechar_motivo = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, ".//*[contains(text(), 'Fechar')]")))
            fechar_motivo.click()
  
  
      try:
          # Escreve "OK" na coluna B se a verificação foi bem sucedida
          if sheet.cell(row=a, column=2).value is None:
            sheet.cell(row=a, column=2).value = "Desfeito"
      except Exception as e:
          print(f"Erro ao escrever na planilha: {e}")
      
      # Salva as alterações na planilha
      workbook.save('TESTE1.xlsx')
      # Navega para a próxima linha da planilha


# Salva as alterações na planilha
workbook.save('TESTE1.xlsx')

# Fecha o navegador
driver.quit()
