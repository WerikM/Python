import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import time
from openpyxl.styles import numbers




# Inicializa o driver do Selenium (no exemplo, será usado o Edge)
driver = webdriver.Edge()

# Abre o site do Zeta
driver.get("SITE")

# seleciona o arquivo da planilha
root = Tk()
root.withdraw()
file_path = askopenfilename()

# abrir a planilha
workbook = openpyxl.load_workbook(file_path)

# selecionar a planilha desejada
try:
    sheet = workbook['Plan1']
except KeyError:
    # caso não encontre, tentar selecionar a planilha com nome "Planilha1"
    try:
        sheet = workbook['Planilha1']
    except KeyError:
        # se não encontrar nenhuma das duas, mostra uma mensagem de erro e sai do programa
        print("Nenhuma planilha encontrada com os nomes 'Plan1' ou 'Planilha1'.")
        exit()

# Percorre as células da coluna A e realiza a navegação para cada valor
for a in range(2, sheet.max_row+1):
    # Obtém a pesquisa a partir da coluna A da linha atual
    termo_pesquisa = str(sheet.cell(row=a, column=2).value)
    print(f"Buscando por: {termo_pesquisa}")
    print(f"pesquisados: {a}")
    if sheet.cell(row=a, column=3).value is None:
        entrega = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div/div/div/div/div[1]/div/div/div[1]/input")))
        entrega.send_keys(termo_pesquisa)
       # time.sleep(2)
        Consultar = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div/div/div/div/div[1]/center/button")))
        Consultar.click()
        try:
            ItemTerceiro = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/section/div/table/tbody[5]/tr/td[2]")))
            texto_Terceiro = ItemTerceiro.text
            # Encontra o elemento pai
            elemento_pai = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/section/div/table/tbody[5]")))
    
            # Verifica o número de filhos do elemento pai
            num_filhos = len(elemento_pai.find_elements(By.XPATH, "./*"))
        except:
           célula = sheet.cell(row=a, column=3)  # especifica a célula usando o número da linha
           célula.value = ("Erro 1")
           workbook.save(file_path)
           FimConsultar = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/a")))
           FimConsultar.click()
        else:
           if num_filhos == 1:
               
               célulaitemTerceiro = sheet.cell(row=a, column=3)  # especifica a célula usando o número da linha
               célulaitemTerceiro.value = texto_Terceiro
               célulaitemTerceiro.number_format = numbers.FORMAT_NUMBER
    
               Cnpj = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/section/div/table/tbody[2]/tr/td[1]")))
               célulaCnpj = sheet.cell(row=a, column=1)  # especifica a célula usando o número da linha
               célulaCnpj.value = Cnpj.text
               célulaCnpj.number_format = numbers.FORMAT_NUMBER
               Qtde = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/section/div/table/tbody[4]/tr[1]/td[1]")))
               célulaQtde = sheet.cell(row=a, column=4)  # especifica a célula usando o número da linha
               célulaQtde.value = Qtde.text
               # Definir o formato da célula como número
               célulaQtde.number_format = numbers.FORMAT_NUMBER
               nome = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/section/div/table/tbody[5]/tr[1]/td[1]")))
               célulanome = sheet.cell(row=a, column=5)  # especifica a célula usando o número da linha
               célulanome.value = nome.text
    
    
        
               FimConsultar = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/a")))
               FimConsultar.click()
               workbook.save(file_path)
           if num_filhos > 1:
               célula = sheet.cell(row=a, column=3)  # especifica a célula usando o número da linha
               célula.value = ("Muitas entregas")
               FimConsultar = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[2]/a")))
               FimConsultar.click()
               workbook.save(file_path)




workbook.save(file_path)





# Fechar o navegador após um tempo (opcional)
# driver.quit()
